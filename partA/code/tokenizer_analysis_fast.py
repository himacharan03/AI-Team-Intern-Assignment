#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

import tiktoken
from openpyxl import load_workbook
from transformers import AutoTokenizer

BASE_DIR = Path(__file__).resolve().parents[2]
WORKBOOK = BASE_DIR / "FLORES_200_4_Languages.xlsx"
OUT_DIR = BASE_DIR / "partA" / "analysis"


def load_rows(limit_per_lang: int = 200):
    wb = load_workbook(WORKBOOK, read_only=True)
    rows = list(wb["All_Sentences"].iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}
    by_lang = defaultdict(list)
    for row in rows[1:]:
        if row[idx["split"]] != "dev":
            continue
        lang = row[idx["language"]]
        if len(by_lang[lang]) >= limit_per_lang:
            continue
        by_lang[lang].append(row[idx["sentence"]])
    return by_lang


def count_words(text: str) -> int:
    return len(re.findall(r"\S+", text, flags=re.UNICODE))


def count_utf8_bytes(text: str) -> int:
    return len(text.encode("utf-8"))


def tokenize_rows(rows, encode):
    tokens_total = 0
    words_total = 0
    bytes_total = 0
    sentence_count = 0
    for text in rows:
        if text is None:
            continue
        tokens = encode(text)
        tokens_total += len(tokens)
        words_total += count_words(text)
        bytes_total += count_utf8_bytes(text)
        sentence_count += 1
    return {
        "sentences": sentence_count,
        "tokens_per_word": (tokens_total / words_total) if words_total else 0.0,
        "tokens_per_utf8_byte": (tokens_total / bytes_total) if bytes_total else 0.0,
        "tokens_per_sentence": (tokens_total / sentence_count) if sentence_count else 0.0,
    }


def main():
    data = load_rows(limit_per_lang=200)
    xlm_tokenizer = AutoTokenizer.from_pretrained("xlm-roberta-base", local_files_only=True)
    tokenizers = {
        "gpt2": tiktoken.get_encoding("gpt2").encode,
        "xlm-roberta-base": lambda s: xlm_tokenizer.encode(s, add_special_tokens=False),
    }
    rows = []
    for tokenizer_name, encode in tokenizers.items():
        for lang in ["English", "Hindi", "Kannada", "Tamil"]:
            stats = tokenize_rows(data.get(lang, []), encode)
            rows.append({
                "tokenizer": tokenizer_name,
                "language": lang,
                "sentences": stats["sentences"],
                "tokens_per_word": stats["tokens_per_word"],
                "tokens_per_utf8_byte": stats["tokens_per_utf8_byte"],
                "tokens_per_sentence": stats["tokens_per_sentence"],
            })
    out = OUT_DIR / "tokenizer_results.csv"
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["tokenizer","language","sentences","tokens_per_word","tokens_per_utf8_byte","tokens_per_sentence"])
        writer.writeheader()
        writer.writerows(rows)
    print(out)
    for r in rows:
        print(r)


if __name__ == "__main__":
    main()
