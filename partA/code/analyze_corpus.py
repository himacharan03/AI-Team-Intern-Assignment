#!/usr/bin/env python3
from __future__ import annotations

import csv
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[2]
WORKBOOK = BASE_DIR / "FLORES_200_4_Languages.xlsx"
OUT_DIR = BASE_DIR / "partA" / "corpus"


def compute_stats():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws_all = wb["All_Sentences"]
    rows = list(ws_all.iter_rows(values_only=True))
    header = rows[0]
    data = rows[1:]

    lang_counts = Counter(r[2] for r in data if r[2] is not None)
    split_counts = Counter(r[0] for r in data if r[0] is not None)
    empty_count = sum(1 for r in data if r[4] is None or str(r[4]).strip() == "")
    duplicates = sum(v - 1 for v in Counter(r[4] for r in data if r[4] is not None).values() if v > 1)

    by_lang = {}
    for lang in ["English", "Hindi", "Kannada", "Tamil"]:
        lang_rows = [r for r in data if r[2] == lang]
        chars = [len(str(r[4])) for r in lang_rows if r[4] is not None]
        words = [len(str(r[4]).split()) for r in lang_rows if r[4] is not None]
        by_lang[lang] = {
            "total_rows": len(lang_rows),
            "avg_chars": sum(chars) / len(chars) if chars else 0,
            "avg_words": sum(words) / len(words) if words else 0,
            "min_chars": min(chars) if chars else 0,
            "max_chars": max(chars) if chars else 0,
            "empty_rows": sum(1 for r in lang_rows if r[4] is None or str(r[4]).strip() == ""),
            "duplicate_rows": 0,
        }

    all_sentences = [str(r[4]) for r in data if r[4] is not None]
    all_chars = [len(text) for text in all_sentences]
    all_words = [len(text.split()) for text in all_sentences]
    summary = {
        "All_Sentences": {
            "total_rows": len(data),
            "split_counts": dict(split_counts),
            "avg_chars": sum(all_chars) / len(all_chars),
            "avg_words": sum(all_words) / len(all_words),
            "min_chars": min(all_chars),
            "max_chars": max(all_chars),
            "empty_rows": empty_count,
            "duplicate_rows": duplicates,
        }
    }
    for lang, stats in by_lang.items():
        summary[lang] = stats

    with (OUT_DIR / "corpus_statistics.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "language",
            "total_rows",
            "split_counts",
            "avg_chars_per_sentence",
            "avg_whitespace_words_per_sentence",
            "min_chars",
            "max_chars",
            "empty_rows",
            "duplicate_rows",
        ])
        for lang, stats in summary.items():
            writer.writerow([
                lang,
                stats["total_rows"],
                stats.get("split_counts", "n/a"),
                stats.get("avg_chars", 0),
                stats.get("avg_words", 0),
                stats.get("min_chars", 0),
                stats.get("max_chars", 0),
                stats.get("empty_rows", 0),
                stats.get("duplicate_rows", 0),
            ])

    print("WORKBOOK:", WORKBOOK)
    print("HEADER:", header)
    print("LANG_COUNTS:", dict(lang_counts))
    print("SPLIT_COUNTS:", dict(split_counts))
    print("EMPTY_ROWS:", empty_count)
    print("DUPLICATE_SENTS:", duplicates)
    print("PER_LANGUAGE_STATS:")
    for lang, stats in by_lang.items():
        print(lang, stats)


if __name__ == "__main__":
    compute_stats()
