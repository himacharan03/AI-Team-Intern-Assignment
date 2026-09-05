#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from pathlib import Path

import tiktoken
from openpyxl import load_workbook
from transformers import AutoTokenizer

BASE_DIR = Path(__file__).resolve().parents[2]
WORKBOOK = BASE_DIR / "FLORES_200_4_Languages.xlsx"
OUT_DIR = BASE_DIR / "partA" / "analysis"


def load_parallel_dev_rows():
    wb = load_workbook(WORKBOOK, read_only=True)
    ws = wb["All_Sentences"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {name: i for i, name in enumerate(header)}
    data = {}
    for row in rows[1:]:
        if row[idx["split"]] != "dev":
            continue
        lang = row[idx["language"]]
        sent = row[idx["sentence"]]
        if lang not in data:
            data[lang] = []
        data[lang].append(sent)
    return data


def byte_count(text: str) -> int:
    return len(text.encode("utf-8"))


def word_count(text: str) -> int:
    text = text.strip()
    if not text:
        return 0
    return len(re.findall(r"\S+", text, flags=re.UNICODE))


def sentence_tokens(tokens):
    return len(tokens)


def analyze_language(lang_texts, tokenizer_name, encode_fn):
    totals = {"tokens": 0, "word_tokens": 0, "bytes": 0, "words": 0, "sentences": 0}
    for text in lang_texts:
        if text is None:
            continue
        tokens = encode_fn(text)
        totals["tokens"] += len(tokens)
        totals["word_tokens"] += len(tokens)
        totals["bytes"] += byte_count(text)
        totals["words"] += word_count(text)
        totals["sentences"] += 1

    return {
        "tokenizer": tokenizer_name,
        "sentences": totals["sentences"],
        "avg_tokens_per_word": totals["tokens"] / totals["words"] if totals["words"] else 0.0,
        "avg_tokens_per_utf8_byte": totals["tokens"] / totals["bytes"] if totals["bytes"] else 0.0,
        "avg_tokens_per_sentence": totals["tokens"] / totals["sentences"] if totals["sentences"] else 0.0,
    }


def build_results():
    data = load_parallel_dev_rows()
    xlm_tokenizer = AutoTokenizer.from_pretrained(
        "xlm-roberta-base",
        local_files_only=True,
    )
    tokenizers = {
        "gpt2": lambda s: tiktoken.get_encoding("gpt2").encode(s),
        "xlm-roberta-base": lambda s: xlm_tokenizer.encode(s, add_special_tokens=False),
    }
    results = {}
    for name, fn in tokenizers.items():
        results[name] = {}
        for lang, texts in data.items():
            results[name][lang] = analyze_language(texts, name, fn)
    return results


def write_csv(results):
    rows = []
    for tokenizer_name, by_lang in results.items():
        for lang, stats in by_lang.items():
            rows.append({
                "tokenizer": tokenizer_name,
                "language": lang,
                "sentences": stats["sentences"],
                "avg_tokens_per_word": stats["avg_tokens_per_word"],
                "avg_tokens_per_utf8_byte": stats["avg_tokens_per_utf8_byte"],
                "avg_tokens_per_sentence": stats["avg_tokens_per_sentence"],
            })

    out_path = OUT_DIR / "tokenizer_results_full_dev.csv"
    with out_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["tokenizer", "language", "sentences", "avg_tokens_per_word", "avg_tokens_per_utf8_byte", "avg_tokens_per_sentence"])
        writer.writeheader()
        writer.writerows(rows)
    print(f"WROTE {out_path}")
    return out_path


if __name__ == "__main__":
    results = build_results()
    for tokenizer_name, by_lang in results.items():
        print(f"TOKENIZER: {tokenizer_name}")
        for lang, stats in by_lang.items():
            print(lang, json.dumps(stats, indent=2, ensure_ascii=False))
    write_csv(results)
